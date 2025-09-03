import pandas as pd
import openpyxl

# Read the Excel file
file_path = r"c:\Users\donpa\Desktop\happy_home\База недвижимости Петропавловск.xlsx"

# Load the workbook to see all sheet names
wb = openpyxl.load_workbook(file_path)
print("Листы в файле:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

print("\n" + "="*80 + "\n")

# Read all sheets
for sheet_name in wb.sheetnames:
    print(f"ЛИСТ: {sheet_name}")
    print("-" * 50)
    
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        print(f"Количество строк: {len(df)}")
        print(f"Количество столбцов: {len(df.columns)}")
        print(f"Столбцы: {list(df.columns)}")
        
        print("\nПервые 3 строки:")
        print(df.head(3).to_string())
        
        print("\nУникальные значения по столбцам:")
        for col in df.columns:
            unique_values = df[col].dropna().unique()
            if len(unique_values) <= 50:  # Show only if not too many values
                print(f"\n{col}:")
                for val in sorted(unique_values, key=str):
                    print(f"  - {val}")
            else:
                print(f"\n{col}: {len(unique_values)} уникальных значений (слишком много для отображения)")
                # Show some examples
                sample_values = sorted(unique_values, key=str)[:10]
                print(f"  Примеры: {sample_values}")
        
    except Exception as e:
        print(f"Ошибка при чтении листа {sheet_name}: {e}")
    
    print("\n" + "="*80 + "\n")